"""Excel export builder for conversation data.

Query strategy: batch by conversation IDs to avoid a single giant JOIN.
Keeps memory bounded and lets the DB use indexes efficiently.
"""
import io
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from sqlalchemy import func, text
from sqlalchemy.orm import Session

from backend.core.enums import AppointmentStatus, FollowupStatus
from backend.core.channel_types import CHANNEL_DISPLAY_NAMES
from backend.models.agent import Agent
from backend.models.appointment import Appointment
from backend.models.conversation import Conversation
from backend.models.conversation_summary import ConversationSummary
from backend.models.message import Message
from backend.models.scheduled_followup import ScheduledFollowup
from backend.models.user import User

MAX_EXPORT_CONVERSATIONS = 4000
_BATCH_SIZE = 100

_HEADER_FONT = Font(bold=True, color="FFFFFF")
_HEADER_FILL = PatternFill("solid", fgColor="3B0764")


def build_export(
    db: Session,
    agent_id: int,
    from_date: str,
    to_date: str,
) -> tuple[io.BytesIO, str]:
    """Build and return Excel workbook as BytesIO + filename.

    Raises ValueError for invalid agent or over-limit results.
    """
    agent = db.get(Agent, agent_id)
    if not agent:
        raise ValueError("Agent not found")

    start_dt, end_dt = _parse_date_range(from_date, to_date)
    conv_rows = _fetch_conversations(db, agent_id, start_dt, end_dt)

    if len(conv_rows) > MAX_EXPORT_CONVERSATIONS:
        raise ValueError(
            f"יש {len(conv_rows):,} שיחות בטווח זה. "
            f"ניתן לייצא עד {MAX_EXPORT_CONVERSATIONS:,} שיחות בפעם אחת."
        )

    wb = Workbook()
    channel_stats = _fetch_channel_stats(db, agent_id, start_dt, end_dt)
    _build_summary_sheet(wb, agent.name, from_date, to_date, len(conv_rows), channel_stats)
    _build_conversations_sheet(wb, db, agent.name, conv_rows)
    _build_channel_stats_sheet(wb, channel_stats)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer, _make_filename(agent.name, from_date, to_date)


# ── Date helpers ──────────────────────────────────────────────────────────────


def _parse_date_range(from_date: str, to_date: str) -> tuple[datetime, datetime]:
    start = datetime.strptime(from_date, "%Y-%m-%d")
    end = datetime.strptime(to_date, "%Y-%m-%d").replace(hour=23, minute=59, second=59)
    return start, end


def _fmt_date(iso: str) -> str:
    return datetime.strptime(iso, "%Y-%m-%d").strftime("%d/%m/%Y")


def _make_filename(agent_name: str, from_date: str, to_date: str) -> str:
    safe = "".join(c if (c.isascii() and c.isalnum()) or c in "-_" else "_" for c in agent_name)
    return f"conversations_{safe}_{from_date}_to_{to_date}.xlsx"


# ── DB queries ────────────────────────────────────────────────────────────────


def _fetch_conversations(
    db: Session, agent_id: int, start_dt: datetime, end_dt: datetime
) -> list[tuple]:
    """Return (conv_id, created_at, user_id, channel_type, channel_external_id)."""
    rows = db.execute(text("""
        SELECT c.id,
               c.created_at,
               c.user_id,
               c.channel_type_snapshot AS channel_type,
               COALESCE(cu.external_id, u.phone) AS customer_id
        FROM conversations c
        JOIN users u ON u.id = c.user_id
        LEFT JOIN channel_users cu ON cu.id = c.channel_user_id
        WHERE c.agent_id = :agent_id
          AND c.created_at >= :start_dt
          AND c.created_at <= :end_dt
        ORDER BY c.created_at
        LIMIT :lim
    """), {
        "agent_id": agent_id,
        "start_dt": start_dt,
        "end_dt": end_dt,
        "lim": MAX_EXPORT_CONVERSATIONS + 1,
    }).fetchall()
    return rows


def _fetch_channel_stats(
    db: Session, agent_id: int, start_dt: datetime, end_dt: datetime
) -> list[dict]:
    """Channel-level stats for Sheet 3."""
    rows = db.execute(text("""
        SELECT
            COALESCE(c.channel_type_snapshot, 'legacy') AS channel_type,
            COUNT(DISTINCT c.id)                          AS conversations,
            COUNT(m.id)                                   AS messages,
            COUNT(DISTINCT CASE WHEN a.id IS NOT NULL THEN c.id END) AS with_appointment
        FROM conversations c
        LEFT JOIN messages m ON m.conversation_id = c.id AND m.role = 'user'
            AND m.created_at >= :start_dt AND m.created_at <= :end_dt
        LEFT JOIN appointments a ON a.conversation_id = c.id
        WHERE c.agent_id = :agent_id
          AND c.created_at >= :start_dt
          AND c.created_at <= :end_dt
        GROUP BY channel_type
        ORDER BY conversations DESC
    """), {"agent_id": agent_id, "start_dt": start_dt, "end_dt": end_dt}).fetchall()

    result = []
    for r in rows:
        convs = int(r.conversations or 0)
        msgs = int(r.messages or 0)
        result.append({
            "channel_type": r.channel_type,
            "display_name": CHANNEL_DISPLAY_NAMES.get(r.channel_type, r.channel_type),
            "conversations": convs,
            "messages": msgs,
            "avg_msg_per_conv": round(msgs / convs, 1) if convs else 0,
            "with_appointment": int(r.with_appointment or 0),
        })
    return result


# ── Sheet builders ────────────────────────────────────────────────────────────


def _build_summary_sheet(
    wb: Workbook,
    agent_name: str,
    from_date: str,
    to_date: str,
    total: int,
    channel_stats: list[dict],
) -> None:
    ws = wb.active
    ws.title = "Summary"

    rows = [
        ("Agent", agent_name),
        ("Date Range", f"{_fmt_date(from_date)} – {_fmt_date(to_date)}"),
        ("Total Conversations", total),
        ("Exported At", datetime.utcnow().strftime("%Y-%m-%d %H:%M UTC")),
        ("", ""),
        ("Channel Breakdown", ""),
    ]
    for i, (label, value) in enumerate(rows, start=1):
        cell = ws.cell(row=i, column=1, value=label)
        cell.font = Font(bold=True)
        ws.cell(row=i, column=2, value=value)

    offset = len(rows) + 1
    ws.cell(row=offset, column=1, value="Channel").font = Font(bold=True)
    ws.cell(row=offset, column=2, value="Conversations").font = Font(bold=True)
    ws.cell(row=offset, column=3, value="Messages").font = Font(bold=True)
    for j, stat in enumerate(channel_stats, start=offset + 1):
        ws.cell(row=j, column=1, value=stat["display_name"])
        ws.cell(row=j, column=2, value=stat["conversations"])
        ws.cell(row=j, column=3, value=stat["messages"])

    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 15


def _build_conversations_sheet(
    wb: Workbook,
    db: Session,
    agent_name: str,
    conv_rows: list[tuple],
) -> None:
    ws = wb.create_sheet("Conversations")
    _write_header(ws)

    user_cache: dict[int, User] = {}

    for batch_start in range(0, len(conv_rows), _BATCH_SIZE):
        batch = conv_rows[batch_start : batch_start + _BATCH_SIZE]
        _write_batch(ws, db, agent_name, batch, batch_start, user_cache)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = "A1:J1"


def _build_channel_stats_sheet(wb: Workbook, channel_stats: list[dict]) -> None:
    """Sheet 3: Channel Stats summary."""
    ws = wb.create_sheet("Channel Stats")
    headers = ["Channel", "Conversations", "Messages", "Avg Msgs/Conv", "With Appointment"]
    col_widths = [24, 16, 12, 16, 18]

    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for col, width in enumerate(col_widths, start=1):
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = width

    for i, stat in enumerate(channel_stats, start=2):
        ws.cell(row=i, column=1, value=stat["display_name"])
        ws.cell(row=i, column=2, value=stat["conversations"])
        ws.cell(row=i, column=3, value=stat["messages"])
        ws.cell(row=i, column=4, value=stat["avg_msg_per_conv"])
        ws.cell(row=i, column=5, value=stat["with_appointment"])


def _write_header(ws) -> None:
    headers = [
        "Agent", "Customer Name", "Customer ID", "Channel", "Phone", "Opened At",
        "Messages", "Appointment Booked", "Conversation", "AI Summary",
    ]
    col_widths = [20, 20, 22, 18, 18, 20, 10, 18, 70, 40]

    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for col, width in enumerate(col_widths, start=1):
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = width


def _write_batch(
    ws,
    db: Session,
    agent_name: str,
    batch: list[tuple],
    batch_start: int,
    user_cache: dict[int, User],
) -> None:
    conv_ids = [r[0] for r in batch]
    user_ids = list({r[2] for r in batch})

    _cache_users(db, user_ids, user_cache)
    messages_by_conv = _load_messages(db, conv_ids)
    followups_by_conv = _load_followups(db, conv_ids)
    summaries_by_conv = _load_summaries(db, conv_ids)
    users_with_appointments = _load_appointment_user_ids(db, user_ids)

    for offset, row in enumerate(batch):
        conv_id = row[0]
        created_at = row[1]
        user_id = row[2]
        channel_type = row[3] if len(row) > 3 else None
        customer_id = row[4] if len(row) > 4 else None

        row_idx = batch_start + offset + 2  # +1 header, +1 1-indexed

        user = user_cache.get(user_id)
        msgs = messages_by_conv.get(conv_id, [])
        followups = followups_by_conv.get(conv_id, [])
        channel_display = CHANNEL_DISPLAY_NAMES.get(channel_type, channel_type or "legacy")

        data = [
            agent_name,
            (user.name or "") if user else "",
            customer_id or (user.phone if user else ""),
            channel_display,
            user.phone if user else "",
            created_at.strftime("%Y-%m-%d %H:%M"),
            len(msgs),
            "Yes" if user_id in users_with_appointments else "No",
            _format_conversation(msgs, followups),
            summaries_by_conv.get(conv_id, ""),
        ]
        for col, value in enumerate(data, start=1):
            cell = ws.cell(row=row_idx, column=col, value=value)
            wrap = col in (9, 10)
            cell.alignment = Alignment(wrap_text=wrap, vertical="top")


# ── Bulk loaders ──────────────────────────────────────────────────────────────


def _cache_users(db: Session, user_ids: list[int], cache: dict[int, User]) -> None:
    missing = [uid for uid in user_ids if uid not in cache]
    if missing:
        for u in db.query(User).filter(User.id.in_(missing)).all():
            cache[u.id] = u


def _load_messages(
    db: Session, conv_ids: list[int]
) -> dict[int, list[Message]]:
    messages = (
        db.query(Message)
        .filter(Message.conversation_id.in_(conv_ids))
        .order_by(Message.conversation_id, Message.created_at)
        .all()
    )
    result: dict[int, list[Message]] = {}
    for msg in messages:
        result.setdefault(msg.conversation_id, []).append(msg)
    return result


def _load_followups(
    db: Session, conv_ids: list[int]
) -> dict[int, list[ScheduledFollowup]]:
    followups = (
        db.query(ScheduledFollowup)
        .filter(
            ScheduledFollowup.conversation_id.in_(conv_ids),
            ScheduledFollowup.status == FollowupStatus.SENT,
            ScheduledFollowup.content.isnot(None),
            ScheduledFollowup.sent_at.isnot(None),
        )
        .order_by(ScheduledFollowup.conversation_id, ScheduledFollowup.sent_at)
        .all()
    )
    result: dict[int, list[ScheduledFollowup]] = {}
    for fu in followups:
        result.setdefault(fu.conversation_id, []).append(fu)
    return result


def _load_summaries(db: Session, conv_ids: list[int]) -> dict[int, str]:
    subq = (
        db.query(
            ConversationSummary.conversation_id,
            func.max(ConversationSummary.created_at).label("max_created"),
        )
        .filter(ConversationSummary.conversation_id.in_(conv_ids))
        .group_by(ConversationSummary.conversation_id)
        .subquery()
    )
    summaries = (
        db.query(ConversationSummary)
        .join(
            subq,
            (ConversationSummary.conversation_id == subq.c.conversation_id)
            & (ConversationSummary.created_at == subq.c.max_created),
        )
        .all()
    )
    return {s.conversation_id: s.summary_text for s in summaries}


def _load_appointment_user_ids(db: Session, user_ids: list[int]) -> set[int]:
    rows = (
        db.query(Appointment.user_id)
        .filter(
            Appointment.user_id.in_(user_ids),
            Appointment.status == AppointmentStatus.SCHEDULED,
        )
        .distinct()
        .all()
    )
    return {row[0] for row in rows}


# ── Conversation formatting ───────────────────────────────────────────────────


def _format_conversation(
    messages: list[Message],
    followups: list[ScheduledFollowup],
) -> str:
    events: list[tuple[datetime, str]] = []

    for msg in messages:
        prefix = "customer" if msg.role == "user" else "agent"
        events.append((msg.created_at, f"{prefix}: {_format_content(msg)}"))

    for fu in followups:
        ts = fu.sent_at.strftime("%Y-%m-%d %H:%M")
        events.append((fu.sent_at, f"--- followup ({ts}) ---\nagent: {fu.content}"))

    events.sort(key=lambda x: x[0])
    return "\n".join(text for _, text in events)


def _format_content(msg: Message) -> str:
    mt = msg.message_type or "text"
    url = msg.media_url or ""

    if mt in ("voice", "audio"):
        return f"[voice message: {url}]" if url else "[voice message]"
    if mt in ("media", "image"):
        return f"[media: {url}]" if url else "[media]"
    return msg.content or ""
